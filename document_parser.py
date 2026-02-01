"""
Document Parser for PDF and Excel files
Handles extraction of text and structured data from various file formats
"""

import os
from typing import List, Dict, Any
from pathlib import Path
import PyPDF2
import openpyxl
from dataclasses import dataclass


@dataclass
class Document:
    """Represents a parsed document with metadata"""
    content: str
    metadata: Dict[str, Any]
    source: str
    doc_type: str


class DocumentParser:
    """Parse PDF and Excel files into structured documents"""

    def __init__(self, files_directory: str = "Files"):
        self.files_directory = files_directory

    def parse_pdf(self, file_path: str) -> Document:
        """
        Extract text from PDF files with metadata

        Args:
            file_path: Path to PDF file

        Returns:
            Document object with extracted text and metadata
        """
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)

                # Extract metadata
                metadata = {
                    'num_pages': len(pdf_reader.pages),
                    'file_name': os.path.basename(file_path),
                    'file_path': file_path
                }

                # Add PDF metadata if available
                if pdf_reader.metadata:
                    metadata.update({
                        'title': pdf_reader.metadata.get('/Title', ''),
                        'author': pdf_reader.metadata.get('/Author', ''),
                        'subject': pdf_reader.metadata.get('/Subject', ''),
                        'creator': pdf_reader.metadata.get('/Creator', '')
                    })

                # Extract text from all pages
                text_content = []
                for page_num, page in enumerate(pdf_reader.pages, 1):
                    page_text = page.extract_text()
                    if page_text.strip():
                        # Add page marker for better context
                        text_content.append(f"[Page {page_num}]\n{page_text}")

                full_text = "\n\n".join(text_content)

                return Document(
                    content=full_text,
                    metadata=metadata,
                    source=file_path,
                    doc_type='pdf'
                )

        except Exception as e:
            print(f"Error parsing PDF {file_path}: {str(e)}")
            raise

    def parse_excel(self, file_path: str) -> Document:
        """
        Extract data from Excel files and convert to structured text

        Args:
            file_path: Path to Excel file

        Returns:
            Document object with extracted data and metadata
        """
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            metadata = {
                'file_name': os.path.basename(file_path),
                'file_path': file_path,
                'sheet_names': workbook.sheetnames,
                'num_sheets': len(workbook.sheetnames)
            }

            # Extract data from all sheets
            text_content = []

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_content.append(f"[Sheet: {sheet_name}]\n")

                # Get all rows with data
                rows_data = []
                for row in sheet.iter_rows(values_only=True):
                    # Filter out completely empty rows
                    if any(cell is not None for cell in row):
                        row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                        rows_data.append(row_text)

                text_content.append("\n".join(rows_data))
                text_content.append("\n")

            full_text = "\n".join(text_content)

            return Document(
                content=full_text,
                metadata=metadata,
                source=file_path,
                doc_type='excel'
            )

        except Exception as e:
            print(f"Error parsing Excel {file_path}: {str(e)}")
            raise

    def parse_all_documents(self) -> List[Document]:
        """
        Parse all PDF and Excel files in the files directory

        Returns:
            List of parsed Document objects
        """
        documents = []
        files_path = Path(self.files_directory)

        if not files_path.exists():
            raise FileNotFoundError(f"Directory '{self.files_directory}' not found")

        # Parse PDF files
        for pdf_file in files_path.glob("*.pdf"):
            print(f"Parsing PDF: {pdf_file.name}")
            doc = self.parse_pdf(str(pdf_file))
            documents.append(doc)

        # Parse Excel files
        for excel_file in files_path.glob("*.xlsx"):
            print(f"Parsing Excel: {excel_file.name}")
            doc = self.parse_excel(str(excel_file))
            documents.append(doc)

        print(f"\nTotal documents parsed: {len(documents)}")
        return documents


if __name__ == "__main__":
    # Test the parser
    parser = DocumentParser()
    docs = parser.parse_all_documents()

    for doc in docs:
        print(f"\n{'='*50}")
        print(f"Source: {doc.source}")
        print(f"Type: {doc.doc_type}")
        print(f"Metadata: {doc.metadata}")
        print(f"Content preview: {doc.content[:200]}...")
